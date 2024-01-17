import csv
import openpyxl
from openpyxl.styles import Font


data_path = r"C:\Users\Gregor Ammann\Science\Results\A028_A029_E005_Pento"

# Some stuff needed to access data in .csv-file
datafile_reader = csv.reader(open(data_path + "\e005_export.csv"))
data = list(datafile_reader)  # 2D-list: data[row][column]


# Font style for emphasis
bold_style = Font(size=11, bold=True)

# Get list of sample names
sample_names = []
for i in range(2, len(data)):
    sample_names.append(data[i][0])

# Get list of modifications
mods = []
mod_concentrations = []
for i in range(len(data[0])):
    cell_content = data[0][i]
    if cell_content.endswith("Method"):
        mod_start_column = i
        mod_name = cell_content[0:-7]
        mods.append(mod_name)
        conc_column = i + 4

        # Get list of modification final concentrations
        concentration_of_modification = []
        for j in range(2, len(data)):
            concentration_of_modification.append(data[j][conc_column])
        mod_concentrations.append(concentration_of_modification)

# Create dictionary with mods as key and concentration list as values
mods_and_conc = {}
for i in range(len(mods)):
    mods_and_conc[mods[i]] = mod_concentrations[i]

# Create new Excel file with two worksheets
wb = openpyxl.Workbook()
wb.active.title = "Mods conc."
sheet1 = wb.active
sheet2 = wb.create_sheet("Additional info")

# Create outline of document: canonicals, mod names, etc
sheet1["B2"] = "Canonicals"
sheet1["B3"] = "C"
sheet1["C3"] = "U"
sheet1["D3"] = "G"
sheet1["E3"] = "A"
sheet1["G3"] = "Sum"
sheet1["B" + str(len(sample_names) + 7)] = "Modifications"
sheet1["B" + str(2 * len(sample_names) + 12)] = "Modification per 1000 canonicals"

for i in range(2, len(mods) + 2):
    sheet1.cell(column=i, row=len(sample_names) + 8).value = mods[i-2]
    sheet1.cell(column=i, row=len(sample_names) + 8).font = bold_style
    sheet1.cell(column=i, row=2 * len(sample_names) + 13).value = mods[i-2]
    sheet1.cell(column=i, row=2 * len(sample_names) + 13).font = bold_style

# Paste mods, conc, and sample names into one
# Sample names
for i in range(4, len(sample_names)+4):
    cell_name = "A" + str(i)
    sheet1[cell_name] = sample_names[i-4]

    cell_name2 = "A" + str(i + len(sample_names) + 5)
    sheet1[cell_name2] = sample_names[i-4]

    cell_name3 = "A" + str(i + 2*len(sample_names) + 10)
    sheet1[cell_name3] = sample_names[i-4]

# Canonical concentrations
for i, conc in enumerate(mods_and_conc["A"]):

    # If a cell is empty, put 0 in there
    if mods_and_conc["A"][i] == "":
        mods_and_conc["A"][i] = 0
    if mods_and_conc["C"][i] == "":
        mods_and_conc["C"][i] = 0
    if mods_and_conc["G"][i] == "":
        mods_and_conc["G"][i] = 0
    if mods_and_conc["U"][i] == "":
        mods_and_conc["U"][i] = 0

    sheet1.cell(column=5, row=i + 4).value = float(mods_and_conc["A"][i])
    sheet1.cell(column=2, row=i + 4).value = float(mods_and_conc["C"][i])
    sheet1.cell(column=4, row=i + 4).value = float(mods_and_conc["G"][i])
    sheet1.cell(column=3, row=i + 4).value = float(mods_and_conc["U"][i])

# Modification concentrations
for i in range(len(mods)):
    mod = sheet1.cell(column=i+2, row=len(sample_names) + 8).value
    for j in range(len(sample_names)):
        if mods_and_conc[mod][j] == "":
            sheet1.cell(column=i + 2, row=j + len(sample_names) + 9).value = 0
        else:
            sheet1.cell(column=i+2, row=j+len(sample_names) + 9).value = float(mods_and_conc[mod][j])


# Calculate sums and per-canonical
# Sum of canonicals
sum_of_canonicals = []
for i in range(len(sample_names)):
    sheet1.cell(row=i+4, column=7).value = float(sheet1.cell(row=i+4, column=2).value) + \
                                           float(sheet1.cell(row=i+4, column=3).value) + \
                                           float(sheet1.cell(row=i+4, column=4).value) + \
                                           float(sheet1.cell(row=i+4, column=5).value)
    sum_of_canonicals.append(sheet1.cell(row=i+4, column=7).value)

# Modification per 1000 canonicals
for i in range(len(mods)):
    mod = sheet1.cell(column=i+2, row=len(sample_names) + 8).value
    for j in range(len(sample_names)):
        if mods_and_conc[mod][j] == "":
            sheet1.cell(column=i + 2, row=2 * len(sample_names) + 14 + j).value = 0
        else:
            if sum_of_canonicals[j] == 0:
                sum_of_canonicals[j] = 0.0001
            sheet1.cell(column=i+2, row=2*len(sample_names)+14+j).value = float(mods_and_conc[mod][j]) / \
                                                                      (sum_of_canonicals[j]/1000)


# Paste additional info into second worksheet
# Get data
mod_count = -1
for i in range(len(data[0])):
    cell_content = data[0][i]
    if cell_content.endswith("Method"):
        mod_start_column = i
        mod_count += 1

        additional_info = []
        for row in range(len(data)):
            one_row = []
            for column in range(mod_start_column, mod_start_column + 6):
                one_row.append(data[row][column])
            additional_info.append(one_row)

        # Paste data
        for info in range(len(additional_info)):
            for item in range(len(additional_info[info])):
                sheet2.cell(row=1 + info + mod_count * (len(sample_names) + 3),
                            column=3 + item).value = additional_info[info][item]


for modification in range(len(mods)):
    # Mod names
    sheet2.cell(row=3 + (modification * (len(sample_names) + 3)), column=1).value = mods[modification]
    sheet2.cell(row=3 + (modification * (len(sample_names) + 3)), column=1).font = bold_style

    # Sample names
    for i in range(len(sample_names)):
        sheet2.cell(3 + (modification * (len(sample_names) + 3) + i), column=2).value = sample_names[i]


# Format text in cells, cell width
# Format cells with text
sheet1["B2"].font = bold_style
sheet1["B3"].font = bold_style
sheet1["C3"].font = bold_style
sheet1["D3"].font = bold_style
sheet1["E3"].font = bold_style
sheet1["G3"].font = bold_style
sheet1["B" + str(len(sample_names) + 7)].font = bold_style
sheet1["B" + str(2 * len(sample_names) + 12)].font = bold_style

# Cell width of col A
sheet1.column_dimensions["A"].width = 30
sheet2.column_dimensions["B"].width = 30

wb.save(data_path + "\\e005.xlsx")
